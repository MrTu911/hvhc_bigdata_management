#!/usr/bin/env python3
"""
Trích xuất dữ liệu thật cán bộ Viện Nghiên cứu Khoa học Hậu cần Quân sự (mã B12)
từ file CSDL.B212.xlsx -> prisma/seed/data/vien_b212_personnel.json.

Nguồn: ~/Downloads/CSDL.B212.xlsx, sheet "DataSheet", header dòng 5-6, data dòng 7-37.
Chuẩn hóa mọi ngày tháng về ISO 'yyyy-mm-dd' (ngày nhập ngũ chỉ có tháng/năm -> ngày 01)
để seed TypeScript dùng trực tiếp new Date(iso).

Chạy: python3 scripts/build_vien_b212_data.py [đường-dẫn-xlsx]
"""
import json
import os
import sys
from datetime import datetime

import openpyxl

# Map tên đơn vị thật -> mã unit trong hệ thống (4 Ban con dưới Viện B12)
UNIT_NAME_TO_CODE = {
    'Chỉ huy Viện Nghiên cứu Khoa học Hậu cần Quân sự': 'B12-CH',
    'Ban Kế hoạch tổng hợp': 'B12-KHTH',
    'Ban Khoa học hậu cần quân sự': 'B12-KHHC',
    'Ban Khoa học kỹ thuật hậu cần': 'B12-KHKT',
}

DEFAULT_XLSX = os.path.expanduser('~/Downloads/CSDL.B212.xlsx')
OUTPUT_PATH = os.path.join(
    os.path.dirname(__file__), '..', 'prisma', 'seed', 'data', 'vien_b212_personnel.json'
)


def normalize_date(value, month_only_day=1):
    """Chuẩn hóa giá trị ngày (datetime hoặc chuỗi) -> ISO 'yyyy-mm-dd' hoặc None.

    Hỗ trợ: datetime; 'dd/mm/yyyy', 'd/m/yyyy'; 'mm/yyyy' (ngày = month_only_day).
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')

    text = str(value).strip()
    if not text:
        return None

    parts = text.split('/')
    try:
        if len(parts) == 3:
            day, month, year = (int(p) for p in parts)
            return datetime(year, month, day).strftime('%Y-%m-%d')
        if len(parts) == 2:
            # Dạng tháng/năm (ngày nhập ngũ) -> dùng ngày mặc định
            month, year = (int(p) for p in parts)
            return datetime(year, month, month_only_day).strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        return None
    return None


def clean_str(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def build_records(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['DataSheet']

    records = []
    unknown_units = set()

    for row in range(7, ws.max_row + 1):
        stt = ws.cell(row=row, column=1).value
        if stt is None:
            continue

        unit_name = clean_str(ws.cell(row=row, column=18).value)
        unit_code = UNIT_NAME_TO_CODE.get(unit_name)
        if unit_name and unit_code is None:
            unknown_units.add(unit_name)

        record = {
            'stt': stt,
            'militaryId': clean_str(ws.cell(row=row, column=2).value),
            'name': clean_str(ws.cell(row=row, column=3).value),
            'dateOfBirth': normalize_date(ws.cell(row=row, column=4).value),
            'gender': clean_str(ws.cell(row=row, column=5).value),
            'birthPlace': clean_str(ws.cell(row=row, column=6).value),
            'placeOfOrigin': clean_str(ws.cell(row=row, column=7).value),
            'ethnicity': clean_str(ws.cell(row=row, column=8).value),
            'religion': clean_str(ws.cell(row=row, column=9).value),
            'bloodGroupRaw': clean_str(ws.cell(row=row, column=10).value),
            'citizenId': clean_str(ws.cell(row=row, column=11).value),
            'citizenIdIssueDate': normalize_date(ws.cell(row=row, column=12).value),
            'citizenIdExpiryDate': normalize_date(ws.cell(row=row, column=13).value),
            'citizenIdIssuePlace': clean_str(ws.cell(row=row, column=14).value),
            'officerIdCard': clean_str(ws.cell(row=row, column=15).value),
            'rank': clean_str(ws.cell(row=row, column=16).value),
            'managementCategoryRaw': clean_str(ws.cell(row=row, column=17).value),
            'unitName': unit_name,
            'unitCode': unit_code,
            'enlistmentDate': normalize_date(ws.cell(row=row, column=19).value),
            'dischargeDate': normalize_date(ws.cell(row=row, column=20).value),
        }
        records.append(record)

    return records, unknown_units


def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(xlsx_path):
        print(f'❌ Không tìm thấy file: {xlsx_path}', file=sys.stderr)
        sys.exit(1)

    records, unknown_units = build_records(xlsx_path)

    if unknown_units:
        print('⚠️  Đơn vị chưa map được sang mã unit:', file=sys.stderr)
        for name in sorted(unknown_units):
            print(f'   - {name}', file=sys.stderr)

    out_path = os.path.abspath(OUTPUT_PATH)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    print(f'✅ Đã ghi {len(records)} bản ghi -> {out_path}')


if __name__ == '__main__':
    main()
